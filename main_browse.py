#! python3
'''
EXAM CREATOR LAUNCHER developed by Mr Steven J walden
    Feb. 2020
    SAMROIYOD, PRACHUAP KIRI KHAN, THAILAND
[See license.txt file]
temp change
'''

__author__ = 'Mr Steven J Walden'
__version__ = '1.1.1'

import os
import sys
import csv
import openpyxl

from app_guis import Ui_CreateCSVWindow

from PyQt5 import QtCore, QtGui, QtWidgets
from PyQt5.QtWidgets import QDesktopWidget
from PyQt5.QtGui import QPalette, QColor
from PyQt5.QtCore import Qt, QT_VERSION_STR

PY_VER = sys.version[:3]

class App(QtWidgets.QWidget):
	"""docstring for App"""
	def __init__(self, parent=None):
		super(App, self).__init__(parent)
		#setup app windows and theme

		self.dark_theme()
		self.load_data()

		self.CreateCSVWindow = QtWidgets.QDialog()
		self.Create_CSV = Ui_CreateCSVWindow()
		#connect the buttons to the methods
		self.CreateCSVWindow.browse_for_workbook = self.browse_for_workbook
		self.CreateCSVWindow.ok_button_clicked = self.ok_button_clicked
		self.CreateCSVWindow.cancel_button_clicked = self.cancel_button_clicked
		self.CreateCSVWindow.populate_sheet_cmb = self.populate_sheet_cmb

		self.Create_CSV.setupUi(self.CreateCSVWindow)

		self.screen_location()

		#limit the number of items in the student name combo box
		self.Create_CSV.SheetcomboBox.setStyleSheet("QComboBox { combobox-popup: 0; }")
		self.Create_CSV.SheetcomboBox.setMaxVisibleItems(10)
		self.Create_CSV.FileBox.setFocus()

		self.CreateCSVWindow.show()

	def load_data(self):
		self.convert_dic = {'A':1,'B':2,'C':3,'D':4,'E':5,'F':6,'G':7,'H':8,'I':9,'J':10,'K':11,'l':12,'M':13,'N':14,'O':15,'P':16,'Q':17,'R':18,'S':19,'T':20,'U':21,'V':22,'W':23,'X':24,'Y':25,'Z':26,'AA':27,'AB':28,'AC':29,'AD':30,'AE':31,'AF':32,'AG':33,'AH':34,'AI':35}

	def dark_theme(self):
		app.setStyle("Fusion")

		self.dark_palette = QPalette()

		self.dark_palette.setColor(QPalette.Window,QColor(53,53,53))
		self.dark_palette.setColor(QPalette.WindowText, Qt.white)
		self.dark_palette.setColor(QPalette.Base, QColor(25, 25, 25))
		self.dark_palette.setColor(QPalette.AlternateBase, QColor(53, 53, 53))
		self.dark_palette.setColor(QPalette.ToolTipBase, Qt.white)
		self.dark_palette.setColor(QPalette.ToolTipText, Qt.white)
		self.dark_palette.setColor(QPalette.Text, Qt.white)
		self.dark_palette.setColor(QPalette.Button, QColor(53, 53, 53))
		self.dark_palette.setColor(QPalette.ButtonText, Qt.white)
		self.dark_palette.setColor(QPalette.BrightText, Qt.red)
		self.dark_palette.setColor(QPalette.Link, QColor(42, 130, 218))
		self.dark_palette.setColor(QPalette.Highlight, QColor(42, 130, 218))
		self.dark_palette.setColor(QPalette.HighlightedText, Qt.black)

		app.setPalette(self.dark_palette)

		app.setStyleSheet("QToolTip { color: #ffffff; background-color: #2a82da; border: 1px solid white; }")

	def screen_location(self):
		ag = QDesktopWidget().availableGeometry()
		sg = QDesktopWidget().screenGeometry()

		self.widget = self.CreateCSVWindow.geometry()
		x = ag.width() / 2 - self.widget.width() / 2
		y = ag.height() / 2 - self.widget.height() / 2
		self.CreateCSVWindow.move(x, y)

	def browse_for_workbook(self):
		self.workbook_name = QtWidgets.QFileDialog.getOpenFileName(None, "Select Excel File", "", "Excel Files (*.xlsx *.xlsm)")
		self.populate_file_box()

	def populate_file_box(self):
		self.Create_CSV.FileBox.setText(self.workbook_name[0])

	def populate_sheet_cmb(self):
		try:
			self.workbook = openpyxl.load_workbook(self.workbook_name[0], data_only = True)
			self.sheet_list = self.workbook.sheetnames
			self.Create_CSV.SheetcomboBox.addItems(self.sheet_list)
		except AttributeError:
			pass

	def copyRange(self, startCol, startRow, endCol, endRow, sheet):
		#Copy range of cells as a nested list
		#Takes: start cell, end cell, and sheet you want to copy from.
	    self.rangeSelected = []
	    #Loops through selected Rows
	    for i in range(startRow,endRow + 1,1):
	        #Appends the row to a RowSelected list
	        self.rowSelected = []
	        for j in range(startCol,endCol+1,1):
	            self.rowSelected.append(sheet.cell(row = i, column = j).value)
	        #Adds the RowSelected List and nests inside the rangeSelected
	        self.rangeSelected.append(self.rowSelected)

	    return self.rangeSelected

	def convert(self, val):
		return self.convert_dic[val]

	def write_csv(self):
		self.read_list =[['Number','Name','Nicknames','Passwords'],['Student number','Choose your name','Nickname','Password']]
		with open('Student_Details_CSV_' + self.sheet_name[-4:] + '.csv', 'w') as new_file:
			csv_writer = csv.writer(new_file, delimiter = ',')

			# csv_writer.writerow(self.read_list)

			for list_detail in self.read_list:
				csv_writer.writerow(list_detail)

			#loop to write the data
			for student_detail in self.rangeSelected:
				csv_writer.writerow(student_detail)

			os.startfile('Student_Details_CSV_' + self.sheet_name[-4:] + '.csv')

	def ok_button_clicked(self):
		try:
			self.sheet_name = self.Create_CSV.SheetcomboBox.currentText()
			self.startcol = self.convert(self.Create_CSV.StartColBox.text())
			self.startrow = int(self.Create_CSV.StartRowBox.text())
			self.endcol = self.convert(self.Create_CSV.EndColBox.text())
			self.endrow = int(self.Create_CSV.EndRowBox.text())

			if self.endcol < self.startcol:
				self.Create_CSV.EndColBox.clear()
			elif self.endrow < self.startrow:
				self.Create_CSV.EndRowBox.clear()
			else:
				self.workbook_sheet = self.workbook[self.sheet_name]
				self.copyRange(self.startcol, self.startrow, self.endcol, self.endrow, self.workbook_sheet)
				self.write_csv()
		except (KeyError, ValueError):
			pass

	def cancel_button_clicked(self):
		self.CreateCSVWindow.close()

print(sys.executable)

if __name__ == '__main__':
    print("Qt version:", QT_VERSION_STR)
    print("Author:", __author__)
    print("App version:",__version__)
    app = QtWidgets.QApplication(sys.argv)
    main_app = App()

sys.exit(app.exec_())
